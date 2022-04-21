import os
import re
from os.path import join, abspath

from openpyxl import load_workbook

TABLE_HEADER = ['ord_det', 'ord', 'det', 'qty', 'part# ', 'qty of p.',
                'material name', 'material type', 'material per 1',
                'unit', 'qty per det']


def get_table_header(ws, min_column=1, max_column=None):
    """
    Return the list of headings from the table
    """
    if not max_column:
        max_column = ws.max_column
    header = [cell.value for cell in next(ws.iter_rows(
        min_row=1, min_col=min_column, max_row=1, max_col=max_column))]
    return header


def imp_table_data(ws_table, min_row=2, min_col=1):
    """
    Import data from a sheet table where the first row is the row ID
    """
    data = {}
    for row in ws_table.iter_rows(min_row=min_row, min_col=min_col, max_row=ws_table.max_row,
                                  max_col=ws_table.max_column, values_only=True):
        key_data = row[0]
        if key_data is not None:
            data[key_data] = [cell for cell in row[1::]]
    return data


def imp_det_list(date: str):
    """
    Import data from a file
    """
    table_name = f'det_list_{date}.xlsx'
    data_path = abspath(join('.', table_name))

    wb = load_workbook(filename=data_path, data_only=True, read_only=True)
    s_name = list(wb.sheetnames)[0]
    ws_t = wb[s_name]
    det_data = imp_table_data(ws_t)

    wb.close()
    return det_data


''' get file list version 2
path_dir_c = join('.', 'calculations')


def file_list(directory):
    filelist = []
    for root, dirs, files in os.walk(directory):
        for filename in files:
            filelist.append(filename)
    return filelist


fl = file_list(path_dir_c)
'''


def file_name_re(name_re, file_list):
    """
    Extract file name according to regular expression
    """
    name_comp = re.compile(name_re)
    for name in file_list:
        name_m = name_comp.match(name)
        if name_m:
            return name_m.group()


def get_files_name(data, path_dir):
    """
    Get file names according to the detail list (calculations)
    """
    list_calc = os.listdir('.' + path_dir)

    det_list_files = {}
    for ord_det in data.keys():
        name_re_calc = str(ord_det + '_r\d{2}_\w{3,4}.xlsx')
        file_list_calc = file_name_re(name_re_calc, list_calc)
        det_list_files[ord_det] = file_list_calc
    return det_list_files


''' get list of files in detail list version 2
def file_name_re_2(name_re, file_list):
    name_f = None
    for name in file_list:
        name_m = re.search(name_re, name)
        if name_m:
            name_f = name_m[0]
    return name_f
'''


def imp_det_calc(ws):
    """
    Get the data from the calculation table
    """
    c_data = []
    '''
    # update table header (make once for TABLE_HEADER)
    table_header_2 = get_table_header(ws)
    table_header += table_header_2[2:5] + table_header_2[9:12]
    table_header.append('qty per det')
    '''
    for row in ws.iter_rows(min_row=2, min_col=2, max_row=ws.max_row, max_col=12, values_only=True):
        if row[0]:
            r1 = row[1]  # check if part #
            if not r1:
                r1 = 1
            qty = int(row[2]) * int(row[9])

            for i in c_data[:]:  # repetition check
                if i[0] == r1:
                    i[1] += row[2]
                    i[6] += qty
                    row = None
            if row:
                c_data_1 = [r1, row[2], row[3], row[8], row[9], row[10], qty]
                c_data.append(c_data_1)
    return c_data


def get_data_calc(det_data, d_list_files, path_dir, s_n):
    """
    Get data from calculation files
    """
    global calc_data
    e_list = []         # detail list without calculation
    for det, file in d_list_files.items():
        if file:
            data_path = abspath(join('.' + path_dir, file))   # path_c
            wb = load_workbook(filename=data_path, data_only=True, read_only=True)
            s_name = list(wb.sheetnames)
            for sn in s_name:
                if sn == s_n:
                    ws = wb[sn]
                    calc_data = imp_det_calc(ws)
            wb.close()
            det_data[det].append(calc_data)
        else:
            e_list.append(det_data.get(det))
    return det_data, e_list
