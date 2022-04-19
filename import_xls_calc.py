"""
Extracting data from multiple xls files and writing to a new one.

This module collects data from files (folder "calculations")
according to the list in the file "det_list _ *. Xlsx"
and combines this data into one file.
"""
from os.path import abspath
from openpyxl import load_workbook, Workbook
from def_xls import *

from openpyxl.formatting import Rule
from openpyxl.styles import Font, PatternFill, Border
from openpyxl.styles.differential import DifferentialStyle

# get details list from xls file
date = '220401'
table_name = f'det_list_{date}.xlsx'
data_path = abspath(join('.', table_name))

wb = load_workbook(filename=data_path, data_only=True, read_only=True)
s_name = list(wb.sheetnames)[0]
ws_t = wb[s_name]

'''
# forming table header (make once for TABLE_HEADER)
table_header = get_table_header(ws_t)
'''

ord_list = []
det_list = []
det_data = {}
for row in ws_t.iter_rows(min_row=2, min_col=1, max_row=ws_t.max_row,
                          max_col=ws_t.max_column, values_only=True):
    key_data = row[0]
    if key_data is not None:
        det_data[key_data] = [cell for cell in row[1::]]
        det_list.append(key_data)
        if row[1] not in ord_list:
            ord_list.append(row[1])

wb.close()

# get file names according to the detail list
path_c = "\\calculations\\"
list_calc = os.listdir('.' + path_c)

det_list_files = {}
for ord_det in det_list:
    name_re_calc = str(ord_det + '_r\d{2}_\w{3,4}.xlsx')
    file_list_calc = file_name_re(name_re_calc, list_calc)
    det_list_files[ord_det] = file_list_calc

# get data from calculations
empty_list = []  # detail list without calculation
for det, file in det_list_files.items():
    calc_data = []
    if file:
        data_path = abspath(join('.' + '\\calculations\\', file))
        wb = load_workbook(filename=data_path, data_only=True, read_only=True)
        s_name = list(wb.sheetnames)
        for sn in s_name:
            if sn == 'матеріали':
                ws = wb[sn]
                '''
                # update table header (make once for TABLE_HEADER)
                table_header_2 = get_table_header(ws)
                table_header += table_header_2[2:5] + table_header_2[9:12]
                table_header.append('qty per det')
                '''
                for row in ws.iter_rows(min_row=2, min_col=2, max_row=ws.max_row, max_col=12, values_only=True):
                    if row[0]:
                        r1 = row[1]  # check if part#
                        if not r1:
                            r1 = 1
                        qty = int(row[2]) * int(row[9])

                        for i in calc_data[:]:      # repetition check
                            if i[0] == r1:
                                i[1] += row[2]
                                i[6] += qty
                                row = None
                        if row:
                            calc_data_1 = [r1, row[2], row[3], row[8], row[9], row[10], qty]
                            calc_data.append(calc_data_1)
        wb.close()
        det_data[det].append(calc_data)
    else:
        empty_list.append(det_data.get(det))


# if the part has a semi-finished product, we must replace the material with it
# from the directory of semi-finished products get file names according to the order list
path_hf = '\\halffabricat_orders\\'
list_hf_files = os.listdir('.' + path_hf)

# select the files that are in the order list
list_hf_files_d = []
for f in list_hf_files:
    d = f.removesuffix('.xlsx')
    if int(d) in ord_list:
        list_hf_files_d.append(f)

# get data from semi-finished products
hf_data = {}
for file in list_hf_files_d:
    data_path = abspath(join('.' + '\\halffabricat_orders\\', file))
    wb = load_workbook(filename=data_path, data_only=True, read_only=True)
    s_name = list(wb.sheetnames)
    for sn in s_name:
        if sn == 'Аркуш1':
            ws = wb[sn]
            for row in ws.iter_rows(min_row=4, min_col=1, max_row=ws.max_row, max_col=5, values_only=True):
                if row[0]:
                    r1 = str(row[0]) + '_' + str(row[1])
                    if r1 not in hf_data.keys():
                        hf_data[r1] = []
                    hf_data[r1].append([row[2], row[4]])

# replace the material with semi-finished product
for hf in hf_data.keys():
    if hf in det_data.keys():
        for hf_val in hf_data.get(hf):
            for v in det_data.get(hf)[3]:
                if hf_val[0] == v[0]:
                    v[3], v[6] = 'halffabricat', int(hf_val[1] / det_data.get(hf)[2])


# create and fill calculations data file
wb = Workbook()
ws = wb.active
ws.append(TABLE_HEADER)

for det in det_data:
    if len(det_data[det]) > 3:
        l = len(det_data[det][3])
        for i in range(0, l):
            row = [det]
            if i > 0:
                row = ['_']
            for r in det_data[det][0:3]:
                row.append(r)
            for c in det_data[det][3][i]:
                row.append(c)
            ws.append(row)

if empty_list:
    ws2 = wb.create_sheet(title='empty')
    ws2['A1'] = 'Without calculation'
    ws2.append(TABLE_HEADER[1:4])
    for d in empty_list:
        ws2.append(d)

    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 18

ws.column_dimensions['A'].width = 28
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 18
ws.column_dimensions['E'].width = 22
ws.column_dimensions['G'].width = 22
ws.column_dimensions['H'].width = 18
ws.column_dimensions['I'].width = 12
ws.column_dimensions['K'].width = 12

dxf = DifferentialStyle(fill=PatternFill(fill_type='solid', start_color='ffff00', end_color='ffff00'))
rule = Rule(operator='equal', type='containsText', text='halffabricat', dxf=dxf)
ws.conditional_formatting.add("H1:H10000", rule)

new_file = f'calculations_{date}.xlsx'
new_file_name = abspath(join('.', 'calculation_data', new_file))
wb.save(new_file_name)
wb.close()
