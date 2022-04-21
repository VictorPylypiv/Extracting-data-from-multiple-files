from os.path import join, abspath

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting import Rule

COLUMN_LIST = ['A', 'B', 'C', 'E', 'G', 'H', 'I', 'K']
WIDTH_LIST = [28, 12, 18, 22, 22, 18, 12, 12]


def col_width(ws, column_list, width_list):
    """
    Set the width of the columns on the sheet
    """
    if len(column_list) == len(width_list):
        for i in range(len(column_list)):
            ws.column_dimensions[column_list[i]].width = width_list[i]


def empty_sheet(list, wb, t_header):
    """
    Return a sheet of details without data based on 'empty list'
    """
    if list:
        ws2 = wb.create_sheet(title='empty')
        ws2['A1'] = 'Without calculation'
        ws2.append(t_header[1:4])
        for d in list:
            ws2.append(d)

        col_width(ws2, COLUMN_LIST, WIDTH_LIST)


def calculation_list(data, ws, col_calculations):
    """
    Return a sheet of details with calculations data based on 'detail data'
    """
    for det in data:
        if len(data[det]) > col_calculations:
            l = len(data[det][col_calculations])
            for i in range(0, l):
                row = [det]
                if i > 0:
                    row = ['_']
                for r in data[det][0:col_calculations]:
                    row.append(r)
                for c in data[det][col_calculations][i]:
                    row.append(c)
                ws.append(row)


def calculation_data(d_data, t_header, date, e_list=None):
    """
    Creating a file filled with calculation data
    """
    wb = Workbook()
    ws = wb.active
    ws.append(t_header)

    calculation_list(d_data, ws, col_calculations=3)

    if e_list:
        empty_sheet(e_list, wb, t_header)

    col_width(ws, COLUMN_LIST, WIDTH_LIST)

    dxf = DifferentialStyle(fill=PatternFill(fill_type='solid', start_color='ffff00', end_color='ffff00'))
    rule = Rule(operator='equal', type='containsText', text='halffabricat', dxf=dxf)
    ws.conditional_formatting.add("H1:H10000", rule)

    new_file = f'calculations_{date}.xlsx'
    new_file_name = abspath(join('.', 'calculation_data', new_file))
    wb.save(new_file_name)
    wb.close()
