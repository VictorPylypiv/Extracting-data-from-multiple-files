import os
from os.path import join, abspath

from openpyxl import load_workbook


def imp_ord_list(data: dict):
    """
    Return list of orders from data
    """
    ord_list = []
    for i in data.values():
        if i[0] not in ord_list:
            ord_list.append(i[0])
    return ord_list


def get_hf_names(data, path_dir):
    """
    Get file names according to the detail list (semi-finished product)
    """
    list_hf_files = os.listdir('.' + path_dir)
    ord_list = imp_ord_list(data)
    list_hf_files_d = []
    for f in list_hf_files:         # select the files that are in the order list
        d = f.removesuffix('.xlsx')
        if int(d) in ord_list:
            list_hf_files_d.append(f)
    return list_hf_files_d


def imp_hf_data(data, path_dir):
    """
    Get data from semi-finished products
    """
    list_hf_files_ = get_hf_names(data, path_dir)
    data = {}
    for file in list_hf_files_:
        data_path = abspath(join('.' + path_dir, file))
        wb = load_workbook(filename=data_path, data_only=True, read_only=True)
        s_name = list(wb.sheetnames)
        for sn in s_name:
            if sn == 'Аркуш1':
                ws = wb[sn]
                for row in ws.iter_rows(min_row=4, min_col=1, max_row=ws.max_row, max_col=5, values_only=True):
                    if row[0]:
                        r1 = str(row[0]) + '_' + str(row[1])
                        if r1 not in data.keys():
                            data[r1] = []
                        data[r1].append([row[2], row[4]])
    return data


def get_data_calc_hf(d_data, path_dir):
    """
    Replace the material with semi-finished product.
    If the part has a semi-finished product, we must replace the material with it
    from the directory of semi-finished products get file names according to the order list.
    """
    h_data = imp_hf_data(d_data, path_dir)
    for hf in h_data.keys():
        if hf in d_data.keys():
            for hf_val in h_data.get(hf):
                # print(1) #
                for v in d_data.get(hf)[3]:
                    if hf_val[0] == v[0]:
                        v[3], v[6] = 'halffabricat', int(hf_val[1] / d_data.get(hf)[2])
    return d_data
